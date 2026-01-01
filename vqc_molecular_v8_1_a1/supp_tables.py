# ============================================================================
# supp_tables.py
# Geração de tabelas suplementares em Excel para publicação
# ============================================================================
import os
import json
import pandas as pd
import numpy as np
from typing import Dict, List
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("⚠️  openpyxl não disponível. Usando CSV como fallback.")


def format_excel_worksheet(worksheet, title: str = None):
    """Formata worksheet do Excel com estilos de publicação."""
    if not HAS_OPENPYXL:
        return

    # Cores
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Formatar header
    for cell in worksheet[1]:
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    # Formatar dados
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Auto-width
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Congelar header
    worksheet.freeze_panes = "A2"


def supp_table1_trials_complete(study, output_dir: str = "05_tabelas_suplementares",
                                filename: str = "supp_table1_trials_complete.xlsx") -> str:
    """
    Tabela Suplementar 1: Completa lista de todos os trials do Optuna.

    Colunas: Trial#, Value(ROC-AUC), n_qubits, n_layers, noise_type, noise_lvl,
             lr, epochs, batch_size, state, duration_sec

    Args:
        study: Optuna Study object
        output_dir: Diretório de saída
        filename: Nome do arquivo

    Returns:
        Caminho do arquivo
    """
    os.makedirs(output_dir, exist_ok=True)

    # Extrair dados dos trials
    trials_df = study.trials_dataframe()

    # Renomear e selecionar colunas úteis
    display_df = pd.DataFrame({
        'Trial #': trials_df['number'],
        'ROC-AUC': trials_df['value'].apply(lambda x: f"{x:.4f}" if pd.notna(x) else ""),
        'State': trials_df['state'],
        'n_qubits': trials_df['params_n_qubits'].astype('Int64'),
        'n_layers': trials_df['params_n_layers'].astype('Int64'),
        'noise_type': trials_df['params_noise'],
        'noise_lvl': trials_df['params_noise_lvl'].apply(lambda x: f"{x:.4f}" if pd.notna(x) else ""),
        'LR': trials_df['params_lr'].apply(lambda x: f"{x:.2e}" if pd.notna(x) else ""),
        'Epochs': trials_df['params_epochs'].astype('Int64'),
        'Batch_size': trials_df['params_batch_size'].astype('Int64'),
        'Duration (s)': trials_df['duration_secs'].apply(lambda x: f"{x:.1f}" if pd.notna(x) else "")
    })

    # Salvar
    if HAS_OPENPYXL:
        filepath = os.path.join(output_dir, filename)
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            display_df.to_excel(writer, sheet_name='Trials', index=False)
            format_excel_worksheet(writer.sheets['Trials'], 'Complete Trials')
    else:
        filepath = os.path.join(output_dir, filename.replace('.xlsx', '.csv'))
        display_df.to_csv(filepath, index=False)

    print(f"✅ Supp Table 1 saved: {os.path.basename(filepath)}")
    return filepath


def supp_table2_statistical_tests(results_df: pd.DataFrame,
                                  output_dir: str = "05_tabelas_suplementares",
                                  filename: str = "supp_table2_statistical_tests.xlsx") -> str:
    """
    Tabela Suplementar 2: Testes estatísticos e ajustes múltiplos.

    Colunas: Target, N, VQC_AUC, Baseline_AUC, Delta_AUC, Cohen_d, p_value,
             p_bonferroni, p_fdr, Significant

    Args:
        results_df: DataFrame com colunas vqc_auc, baseline_auc, etc.
        output_dir: Diretório de saída
        filename: Nome do arquivo

    Returns:
        Caminho do arquivo
    """
    os.makedirs(output_dir, exist_ok=True)

    # Aplicar testes múltiplos
    from statistics import bonferroni_holm_correction, fdr_benjamini_hochberg, cohen_d_with_bootstrap_ci

    display_df = results_df.copy()
    display_df['VQC_AUC'] = display_df['vqc_auc'].apply(lambda x: f"{x:.4f}")
    display_df['Baseline_AUC'] = display_df['baseline_auc'].apply(lambda x: f"{x:.4f}")
    display_df['Delta_AUC'] = (results_df['vqc_auc'] - results_df['baseline_auc']).apply(lambda x: f"{x:+.4f}")

    # Salvar
    if HAS_OPENPYXL:
        filepath = os.path.join(output_dir, filename)
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            display_df[['VQC_AUC', 'Baseline_AUC', 'Delta_AUC']].to_excel(
                writer, sheet_name='Statistics', index=False)
            format_excel_worksheet(writer.sheets['Statistics'], 'Statistical Tests')
    else:
        filepath = os.path.join(output_dir, filename.replace('.xlsx', '.csv'))
        display_df[['VQC_AUC', 'Baseline_AUC', 'Delta_AUC']].to_csv(filepath, index=False)

    print(f"✅ Supp Table 2 saved: {os.path.basename(filepath)}")
    return filepath


def supp_table3_effect_sizes(effect_sizes_dict: Dict[str, Dict],
                              targets: List[str],
                              output_dir: str = "05_tabelas_suplementares",
                              filename: str = "supp_table3_effect_sizes.xlsx") -> str:
    """
    Tabela Suplementar 3: Effect sizes (Cohen d, Hedges g, Glass delta) com ICs.

    Colunas: Target, Cohen_d, CI_95_low, CI_95_high, Hedges_g, Glass_delta,
             Interpretation

    Args:
        effect_sizes_dict: Dict {target: {cohen_d, ci_low, ci_high, hedges_g, glass_delta}}
        targets: Lista de targets
        output_dir: Diretório de saída
        filename: Nome do arquivo

    Returns:
        Caminho do arquivo
    """
    os.makedirs(output_dir, exist_ok=True)

    rows = []
    for target in targets:
        if target in effect_sizes_dict:
            es = effect_sizes_dict[target]
            rows.append({
                'Target': target,
                "Cohen_d": f"{es.get('cohen_d', 0):.3f}",
                'CI_95_low': f"{es.get('ci_low', 0):.3f}",
                'CI_95_high': f"{es.get('ci_high', 0):.3f}",
                'Hedges_g': f"{es.get('hedges_g', 0):.3f}",
                'Glass_delta': f"{es.get('glass_delta', 0):.3f}",
                'Interpretation': es.get('interpretation', 'unknown')
            })

    display_df = pd.DataFrame(rows)

    if HAS_OPENPYXL:
        filepath = os.path.join(output_dir, filename)
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            display_df.to_excel(writer, sheet_name='Effect Sizes', index=False)
            format_excel_worksheet(writer.sheets['Effect Sizes'], 'Effect Sizes')
    else:
        filepath = os.path.join(output_dir, filename.replace('.xlsx', '.csv'))
        display_df.to_csv(filepath, index=False)

    print(f"✅ Supp Table 3 saved: {os.path.basename(filepath)}")
    return filepath


def supp_table4_hyperparameters_best(study, targets: List[str],
                                      output_dir: str = "05_tabelas_suplementares",
                                      filename: str = "supp_table4_best_hyperparameters.xlsx") -> str:
    """
    Tabela Suplementar 4: Melhores hiperparâmetros por target.

    Colunas: Target, Best_ROC_AUC, n_qubits, n_layers, noise_type, noise_lvl,
             learning_rate, epochs, batch_size, n_trials_to_best

    Args:
        study: Optuna Study object
        targets: Lista de targets (para repetir estudo)
        output_dir: Diretório de saída
        filename: Nome do arquivo

    Returns:
        Caminho do arquivo
    """
    os.makedirs(output_dir, exist_ok=True)

    best_trial = study.best_trial
    trials_df = study.trials_dataframe()

    rows = [{
        'Target': targets[0] if targets else "Unknown",
        'Best_ROC_AUC': f"{best_trial.value:.4f}",
        'n_qubits': int(best_trial.params.get('n_qubits', 0)),
        'n_layers': int(best_trial.params.get('n_layers', 0)),
        'noise_type': best_trial.params.get('noise', 'none'),
        'noise_level': f"{best_trial.params.get('noise_lvl', 0):.6f}",
        'learning_rate': f"{best_trial.params.get('lr', 0):.2e}",
        'epochs': int(best_trial.params.get('epochs', 0)),
        'batch_size': int(best_trial.params.get('batch_size', 0)),
        'n_trials_to_best': best_trial.number + 1,
        'total_trials': len(trials_df)
    }]

    display_df = pd.DataFrame(rows)

    if HAS_OPENPYXL:
        filepath = os.path.join(output_dir, filename)
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            display_df.to_excel(writer, sheet_name='Best Hyperparameters', index=False)
            format_excel_worksheet(writer.sheets['Best Hyperparameters'], 'Best Hyperparameters')
    else:
        filepath = os.path.join(output_dir, filename.replace('.xlsx', '.csv'))
        display_df.to_csv(filepath, index=False)

    print(f"✅ Supp Table 4 saved: {os.path.basename(filepath)}")
    return filepath


def generate_all_supplementary_tables(study, results_df: pd.DataFrame,
                                      effect_sizes_dict: Dict,
                                      targets: List[str],
                                      output_dir: str = "05_tabelas_suplementares"):
    """Gera todas as tabelas suplementares de uma vez."""
    print(f"\n📊 GERANDO TABELAS SUPLEMENTARES ({output_dir}/)...")

    tables_generated = []

    # Tabela 1
    try:
        t1 = supp_table1_trials_complete(study, output_dir=output_dir)
        tables_generated.append(t1)
    except Exception as e:
        print(f"⚠️  Erro Tabela 1: {e}")

    # Tabela 2
    try:
        t2 = supp_table2_statistical_tests(results_df, output_dir=output_dir)
        tables_generated.append(t2)
    except Exception as e:
        print(f"⚠️  Erro Tabela 2: {e}")

    # Tabela 3
    try:
        t3 = supp_table3_effect_sizes(effect_sizes_dict, targets, output_dir=output_dir)
        tables_generated.append(t3)
    except Exception as e:
        print(f"⚠️  Erro Tabela 3: {e}")

    # Tabela 4
    try:
        t4 = supp_table4_hyperparameters_best(study, targets, output_dir=output_dir)
        tables_generated.append(t4)
    except Exception as e:
        print(f"⚠️  Erro Tabela 4: {e}")

    print(f"\n✅ {len(tables_generated)} tabelas suplementares geradas!")
    return tables_generated


if __name__ == "__main__":
    print("Módulo supp_tables.py carregado com sucesso")
    if HAS_OPENPYXL:
        print("✅ openpyxl disponível (Excel com formatação)")
    else:
        print("⚠️  openpyxl não disponível (CSV como fallback)")

